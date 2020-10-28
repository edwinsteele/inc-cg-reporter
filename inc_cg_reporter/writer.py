import datetime
from typing import List, Dict

from backports.zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from inc_cg_reporter.connect_group import (
    ConnectGroup,
    ConnectGroupPersonManager,
    Person,
)


class ConnectGroupWorksheetGenerator:
    """Creates a well formatted worksheet for a Connect Group"""

    WIDTH_MULTIPLIER_OVER_FIXED = 1.2
    FIRST_COLUMN_WIDTH = 22
    FIRST_ROW_HEIGHT = 2

    def __init__(self, field_list: List[str]):
        # column indexes start from 1 and enumerate uses zero-based counting,
        #  so we need to bump our column number by one
        self._column_locations = {
            col_name: col_number + 1 for col_number, col_name in (enumerate(field_list))
        }

    def person_as_row_values(self, person: Person) -> Dict[int, str]:
        row = {}
        # XXX this generator shouldn't need to know where to find the personal
        #  attributes on the person.
        for column_name, value in person.personal_attributes.items():
            row[self._column_locations[column_name]] = value

        return row

    def populate(self, ws: Worksheet, cg: ConnectGroup):
        ws.title = cg.name
        self.create_column_headers(ws)
        for person in cg.members:
            ws.append(self.person_as_row_values(person))

    def create_column_headers(self, ws: Worksheet):
        ws.cell(row=1, column=1, value="Name")
        for col_name, col_index in self._column_locations.items():
            ws.cell(row=1, column=col_index, value=col_name)

    def style(self, ws: Worksheet):
        # Size all columns based on name of column (values are all
        #  dates, so won't be the limiting factor)
        for col_name, col_index in self._column_locations.items():
            # column_dimensions requires a column name, not an index
            ws.column_dimensions[get_column_letter(col_index)].width = (
                len(col_name) * self.WIDTH_MULTIPLIER_OVER_FIXED
            )
        # Then override the first column width (it's like a header)
        ws.column_dimensions["A"].width = self.FIRST_COLUMN_WIDTH
        # Style the first column in a header-like way
        for cell in next(iter(ws.iter_cols(min_col=1, max_col=1))):
            cell.style = "40 % - Accent1"

        # Style header row (note the overlap with the name column... we're
        #  intentionally overwriting the style of A1 to be what is below)
        for cell in next(iter(ws.iter_rows(min_row=1, max_row=1))):
            cell.style = "Accent1"


class ConnectGroupWorkbookManager:
    def __init__(
        self,
        cgpm: ConnectGroupPersonManager,
        cgwsg: ConnectGroupWorksheetGenerator,
    ):
        self._cgpm = cgpm
        self._cgwsg = cgwsg
        self._workbook = Workbook()

    def insert_title_sheet(self) -> None:
        about_sheet = self._workbook.create_sheet("About", 0)
        about_sheet.title = "About"
        now_au = datetime.datetime.now(tz=ZoneInfo("Australia/Sydney"))
        about_sheet["A1"] = "Created: {}".format(now_au.ctime())
        about_sheet["A2"] = "Connect Group Count: {}".format(
            self._cgpm.connect_groups_count
        )
        about_sheet["A3"] = "Connect Group Total Member Count: {}".format(
            self._cgpm.connect_groups_member_count
        )

    def create(self) -> None:
        for connect_group in self._cgpm.connect_groups.values():
            ws = self._workbook.create_sheet()
            self._cgwsg.populate(ws, connect_group)
            self._cgwsg.style(ws)

        # Remove the blank sheet that's created initially
        self._workbook.remove(self._workbook.worksheets[0])
        self.insert_title_sheet()

    def save(self, filename: str) -> None:
        self._workbook.save(filename)
