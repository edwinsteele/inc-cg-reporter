import datetime
import pathlib
from tempfile import NamedTemporaryFile
from typing import List, Dict

from zoneinfo import ZoneInfo
from more_itertools import first
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PrintPageSetup
from openpyxl.worksheet.worksheet import Worksheet

from inc_cg_reporter.connect_group import (
    ConnectGroup,
    ConnectGroupMembershipManager,
    Person,
)
from inc_cg_reporter.field_definition import PERSONAL_ATTRIBUTE_NAME


class ConnectGroupWorksheetGenerator:
    """Creates a well formatted worksheet for a Connect Group"""

    DATE_TYPE_COLUMN_WIDTH = 13
    FIRST_COLUMN_WIDTH = 22
    TEAM_COLUMN_WIDTH = 26
    FIRST_ROW_HEIGHT = 2
    HEADER_ROW_HEIGHT = 30
    THIN_BORDER = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin"),
    )

    def __init__(self, field_list: List[str]):
        # column indexes start from 1 and enumerate uses zero-based counting,
        #  so we need to bump our column number by one
        self._column_locations = {
            col_name: col_number + 1 for col_number, col_name in (enumerate(field_list))
        }

    def person_as_row_values(self, person: Person) -> Dict[int, str]:
        row = {}
        # XXX this generator shouldn't need to know where to find the personal
        #  attributes on the person.
        for column_name, value in person.personal_attributes.items():
            row[self._column_locations[column_name]] = value

        return row

    def populate(self, ws: Worksheet, cg: ConnectGroup):
        ws.title = cg.name
        self.create_column_headers(ws)
        for person in sorted(
            cg.members, key=lambda p: p.personal_attributes[PERSONAL_ATTRIBUTE_NAME]
        ):
            ws.append(self.person_as_row_values(person))

    def create_column_headers(self, ws: Worksheet):
        ws.cell(row=1, column=1, value="Name")
        for col_name, col_index in self._column_locations.items():
            ws.cell(row=1, column=col_index, value=col_name)

    def insert_heading(self, ws: Worksheet):
        # This can only be run after the table has been populated and styled
        #  given it prepends rows to the sheet. Ugh.
        ws.insert_rows(0)
        ws["A1"] = ws.title
        ws.merge_cells("A1:{}1".format(get_column_letter(ws.max_column)))
        # Style merged cells using the top left cell reference
        ws["A1"].style = "Headline 1"
        # alignment is overwritten by style, so set it afterwards
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        no_border = Side(border_style=None)
        ws["A1"].border = Border(
            left=no_border, right=no_border, top=no_border, outline=False
        )

    def style(self, ws: Worksheet):
        # Give columns a fixed width so each sheet can print onto a single
        #  A4 in landscape mode.
        for col_name, col_index in self._column_locations.items():
            if col_name == "Team":
                cw = self.TEAM_COLUMN_WIDTH
            else:
                cw = self.DATE_TYPE_COLUMN_WIDTH

            # column_dimensions requires a column name, not an index
            ws.column_dimensions[get_column_letter(col_index)].width = cw

        # Then override the first column width (it's like a header)

        ws.column_dimensions["A"].width = self.FIRST_COLUMN_WIDTH
        # Style the first column in a header-like way
        for cell in first(ws.columns):
            cell.style = "40 % - Accent1"

        # Style header row (note the overlap with the name column... we're
        #  intentionally overwriting the style of A1 to be what is below)
        for cell in first(ws.rows):
            cell.style = "Accent1"
            cell.alignment = Alignment(wrap_text=True, horizontal="center")

        # Intended to be double height, with text wrap set in the loop below
        ws.row_dimensions[1].height = self.HEADER_ROW_HEIGHT

        # Style the data cells (non-header cells)
        for row in ws.iter_rows(min_row=2, min_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
                cell.border = self.THIN_BORDER

    def setup_print_page_setup(self, ws):
        # fitToWidth isn't recognised on numbers. Hardcoding a scale is ghastly,
        #  but it works and will update if it's inappropriate
        ws.page_setup = PrintPageSetup(orientation="landscape", scale=75)


class ConnectGroupWorkbookManager:
    """An excel workbook, with sheets per connect group and a summary sheet"""

    OUTPUT_FILENAME = "inc_cg.xlsx"

    def __init__(
        self,
        membership_manager: ConnectGroupMembershipManager,
        worksheet_generator: ConnectGroupWorksheetGenerator,
    ):
        self._membership_manager = membership_manager
        self._worksheet_generator = worksheet_generator
        self._workbook = Workbook()

    def insert_title_sheet(self) -> None:
        about_sheet = self._workbook.create_sheet("About", 0)
        about_sheet.title = "About"
        about_sheet.column_dimensions["A"].width = 30
        about_sheet.column_dimensions["B"].width = 40
        now_au = datetime.datetime.now(tz=ZoneInfo("Australia/Sydney"))
        about_sheet.append({"A": "Created:", "B": now_au.ctime()})
        about_sheet.append(
            {
                "A": "Connect Group Count:",
                "B": self._membership_manager.connect_groups_count,
            }
        )
        about_sheet.append(
            {
                "A": "Connect Group Total Member Count:",
                "B": self._membership_manager.connect_groups_member_count,
            }
        )
        # Volunteer Counts:
        for (
            volunteer_role,
            role_count,
        ) in self._membership_manager.volunteer_count.items():
            about_sheet.append(
                {
                    "A": f"Team size - {volunteer_role}:",
                    "B": role_count,
                }
            )

        # Show a list of ConnectGroups
        if self._membership_manager.connect_groups_member_count > 0:
            about_sheet.append(
                {"A": "Connect Group List:", "B": self._workbook.worksheets[1].title}
            )
        # Ignore the zeroth worksheet (this about page), and the first worksheet
        #  that we printed in the line about
        for ws in self._workbook.worksheets[2:]:
            about_sheet.append({"B": ws.title})

    def create(self) -> None:
        for connect_group in sorted(
            self._membership_manager.connect_groups.values(), key=lambda x: x.name
        ):
            ws = self._workbook.create_sheet()
            self._worksheet_generator.populate(ws, connect_group)
            self._worksheet_generator.style(ws)
            self._worksheet_generator.insert_heading(ws)
            self._worksheet_generator.setup_print_page_setup(ws)

        # Remove the blank sheet that's created initially
        self._workbook.remove(self._workbook.worksheets[0])
        self.insert_title_sheet()

    def save(self) -> pathlib.Path:
        output_file = NamedTemporaryFile(delete=False)
        self._workbook.save(output_file.name)
        output_file.close()
        output_path = pathlib.Path(output_file.name)
        return output_path.rename(output_path.with_name(self.OUTPUT_FILENAME))
